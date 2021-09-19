import openpyxl
import json
from collections import OrderedDict
import pprint
import time

# エクセルファイルのロード
excel_path = '2021jikanwari_kensaku.xlsx'
workbook = openpyxl.load_workbook(filename=excel_path, read_only=True)

# シートのロード
sheet = workbook['後期']

# max_row_value = sheet.max_row

# 範囲データ取得
sheet_range = sheet['A3':'M1164']

gakubu = []
gakka = []
No = []
niti = []
gen = []
gakki = []
kurasu = []
kamoku = []
tantou = []
kyoushitsu = []
sou = []
ji = []
bikou = []

# print(sheet.max_row)

for row in sheet_range:
    for cell in row:
        # 該当セルの値取得
        get_low_id = cell.coordinate

        if get_low_id[0] == "A":
            gakubu.append(cell.value)
        elif get_low_id[0] == "B":
            gakka.append(cell.value)
        elif get_low_id[0] == "C":
            No.append(cell.value)
        elif get_low_id[0] == "D":
            niti.append(cell.value)
        elif get_low_id[0] == "E":
            gen.append(cell.value)
        elif get_low_id[0] == "F":
            gakki.append(cell.value)
        elif get_low_id[0] == "G":
            kurasu.append(cell.value)
        elif get_low_id[0] == "H":
            kamoku.append(cell.value)
        elif get_low_id[0] == "I":
            tantou.append(cell.value)
        elif get_low_id[0] == "J":
            kyoushitsu.append(cell.value)
        elif get_low_id[0] == "K":
            sou.append(cell.value)
        elif get_low_id[0] == "L":
            ji.append(cell.value)
        elif get_low_id[0] == "M":
            bikou.append(cell.value)

json_list = []
for i in range(len(kamoku)):
    data = OrderedDict()
    data["gakubu"] = gakubu[i]
    data["gakka"] = gakka[i]
    data["No"] = No[i]
    data["niti"] = niti[i]
    data["gen"] = gen[i]
    data["gakki"] = gakki[i]
    data["kurasu"] = kurasu[i]
    data["kamoku"] = kamoku[i]
    data["tantou"] = tantou[i]
    data["kyoushitsu"] = kyoushitsu[i]
    data["sou"] = sou[i]
    data["ji"] = ji[i]
    data["bikou"] = bikou[i]

    json_list.append(data)
output_name = './output/timetable.json'
output_name = output_name.replace('./pdf/', '')
with open(output_name, 'w') as file:
    json.dump(json_list, file, indent=4, ensure_ascii=False)
print("Success!!")
